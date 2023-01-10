from openpyxl import Workbook
from openpyxl import load_workbook
import re
import requests
from bs4 import BeautifulSoup

wb = load_workbook(filename='players2012-2020.xlsx')
ws = wb['PlayersRating']

players = set()
for row in ws.rows:
    players.add((row[0].value, row[1].value))

# ws2 = wb.create_sheet("PlayersRatingOrdered")
# ws2.title = "PlayersRating"

teams = []
for player in players:
    page = requests.get('https://www.hltv.org' + player[0])
    soup = BeautifulSoup(page.content, 'html.parser')
    print(player[0])
    print(soup.prettify())
    teamHtml = soup.find('div', class_ = 'playerTeam')
    if teamHtml is None:
        teamHtml = soup.find('span', class_ = 'profile-player-stat-value bold')
        print("test" + teamHtml.prettify())
    teamName = teamHtml.find('a').get_text()
    teamLink = teamHtml.find('a')['href']
    print(teamName + " " + teamLink)
