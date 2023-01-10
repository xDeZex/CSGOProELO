# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://docs.scrapy.org/en/latest/topics/item-pipeline.html

from openpyxl import Workbook
from openpyxl import load_workbook
import re

class XlsxPipeline(object):
    """Save scraped data to an XLSX spreadsheet"""
    i = 0

    def open_spider(self, spider):
        """Start scraping"""
        # Create an Excel workbook
        self.wb = load_workbook(filename='players2012-2020.xlsx')
        self.ws = self.wb.create_sheet('TeamRatings')
        self.ws.title = "TeamRatings"

    def process_item(self, item, spider):
        # Append a row to the spreadsheet
        if len(item["team"]) == 0:
            return item
        rowNames = []
        rowNames.append(item["teamName"])
        for player in item["team"]:
            rowNames.append(re.search('[^\/]+$', player).group(0))
        if len(rowNames) != 6:
            for x in range(0, 6 - len(rowNames)):
                rowNames.append(None)
        rowNames.append("AVG")
        print(rowNames)
        self.ws.append(rowNames)
        rowScores = []
        rowScores.append("")
        for player in item["teamELO"]:
            rowScores.append(player)
        rowScores.append(sum(item["teamELO"]) / len(item["teamELO"]))
        print(sum(item["teamELO"]) / len(item["teamELO"]))
        self.ws.append(rowScores)
        return item

    def close_spider(self, spider):
        """Stop scraping"""
        # Save the Excel workbook
        self.wb.save('players2012-2020.xlsx')

# useful for handling different item types with a single interface
from itemadapter import ItemAdapter


class TutorialPipeline:
    def process_item(self, item, spider):
        return item
