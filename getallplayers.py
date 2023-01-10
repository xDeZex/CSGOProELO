from openpyxl import Workbook
from openpyxl import load_workbook

def getPlayers(year):
    setPlayers = set()

    wb2 = load_workbook(filename='matches' + year + '.xlsx', read_only=True)
    ws2 = wb2['Matches']

    for row in ws2.rows:
        for cell in row:
            if cell.value:
                if cell.column_letter in ['A', 'B', 'C', 'D', 'E', 'H', 'I', 'J', 'K', 'L']:
                    setPlayers.add(cell.value)
    return setPlayers


wb = Workbook(write_only=True)
wb.title = "Teams"
ws = wb.create_sheet("Players")

ws.title = "Players"

setPlayers = set()



#"2012", "2013", "2014", "2015", "2016", "2017", "2018", "2019", "2020"
for year in ["2012", "2013", "2014", "2015", "2016", "2017", "2018", "2019", "2020", '2021']:
    setPlayers = setPlayers.union(getPlayers(year))


for player in setPlayers:
    ws.append([player, 0])
wb.save('players2012-2020.xlsx')
