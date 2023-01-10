from openpyxl import Workbook
from openpyxl import load_workbook



wb = load_workbook(filename='players2012-2020.xlsx')
ws = wb['Players']


playerElo = {}

for row in ws.rows:
    playerElo[row[0].value] = row[1].value

#"2012", "2013", "2014", "2015", "2016", "2017", "2018", "2019", "2020"
for indexYear, year in enumerate(["2012", "2013", "2014", "2015", "2016", "2017", "2018", "2019", "2020"]):
    wb = load_workbook(filename='matches' + year + '.xlsx')
    ws2 = wb["Matches"]

    for row in ws2.rows:
        team1 = []
        # for i, x in enumerate(row):
        #     print(str(i) + " " + str(x.value))
        for player in row[:5]:
            team1.append(player.value)
        team2 = []
        for player in row[7:12]:
            team2.append(player.value)

        team1 = [player for player in team1 if player is not None]
        team2 = [player for player in team2 if player is not None]

        if len(team1) == 0 or len(team2) == 0:
            continue

        team1Score = row[5].value
        team2Score = row[6].value
        #print(team1 + ": " + str(team1Score) + " - " + str(team2Score) + " :" + team2)
        if team1Score + team2Score > 30:
            if team1Score > team2Score:
                team1Score = 1
                team2Score = 0
            else:
                team1Score = 0
                team2Score = 1
        if team1Score == 16 or team2Score == 16:
            if team1Score == 16:
                team1Score = 1
                team2Score = 0
            else:
                team1Score = 0
                team2Score = 1
        if team1Score == 15 or team2Score == 15:
            team1Score = 0.5
            team2Score = 0.5

        totalMaps = team1Score + team2Score

        team1Rating = [playerElo[player] for player in team1 if playerElo[player] != 0]
        team2Rating = [playerElo[player] for player in team2 if playerElo[player] != 0]

        if len(team1Rating) != 0:
            team1RatingAVG = sum(team1Rating) / len(team1Rating)
        else:
            team1RatingAVG = 0
        if len(team2Rating) != 0:
            team2RatingAVG = sum(team2Rating) / len(team2Rating)
        else:
            team2RatingAVG = 0

        team1playersNoScore = []
        for player in team1:
            if playerElo[player] == 0:
                team1playersNoScore.append(player)
        team2playersNoScore = []
        for player in team2:
            if playerElo[player] == 0:
                team2playersNoScore.append(player)


        for player in team1playersNoScore:
            if team1RatingAVG != 0:
                playerElo[player] = team1RatingAVG
            elif team2RatingAVG != 0:
                playerElo[player] = team2RatingAVG
            else:
                playerElo[player] = 1500 - indexYear * 100
        for player in team2playersNoScore:
            if team2RatingAVG != 0:
                playerElo[player] = team2RatingAVG
            elif team1RatingAVG != 0:
                playerElo[player] = team1RatingAVG
            else:
                playerElo[player] = 1500 - indexYear * 100


        team1Rating = [playerElo[player] for player in team1]
        team2Rating = [playerElo[player] for player in team2]

        if team1RatingAVG == 0:
            team1RatingAVG = sum(team1Rating) / len(team1Rating)
        if team2RatingAVG == 0:
            team2RatingAVG = sum(team2Rating) / len(team2Rating)

        team1ES = 1 / (1 + 10 ** ((team2RatingAVG - team1RatingAVG) / 400)) * totalMaps
        team2ES = 1 / (1 + 10 ** ((team1RatingAVG - team2RatingAVG) / 400)) * totalMaps

        team1newRating = []
        for i, player in enumerate(team1Rating):
            team1newRating.append(player + 32 * (team1Score - team1ES))

        team2newRating = []
        for i, player in enumerate(team2Rating):
            team2newRating.append(player + 32 * (team2Score - team2ES))

        #print(team1 + ": " + str(teamElo[team1]) + " - " + str(teamElo[team2]) + " :" + team2)
        #print(team1 + ": " + str(team1newRating) + " - " + str(team2newRating) + " :" + team2)

        for i, player in enumerate(team1):
            playerElo[player] = team1newRating[i]
        for i, player in enumerate(team2):
            playerElo[player] = team2newRating[i]


wb = load_workbook(filename='players2012-2020.xlsx')
try:
    ws = wb['PlayersRating']
except:
    ws = wb.create_sheet("PlayersRating")
    ws.title = "PlayersRating"

for x, y in playerElo.items():
    ws.append([x, y])

wb.save('players2012-2020.xlsx')
